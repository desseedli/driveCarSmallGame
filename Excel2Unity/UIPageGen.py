import openpyxl
import os
from Config import UIPageLoadXlsxPath
from Config import UnityCodeDir


class UIPageGen:
    def __init__(self):
        self.mData = []

    def Tab(self, count):
        return "    " * count

    def gen_ui_page(self):
        if not os.path.exists(UIPageLoadXlsxPath):
            return False
        workbook = openpyxl.load_workbook(UIPageLoadXlsxPath)
        table = workbook.worksheets[0]
        max_row = table.max_row
        max_column = table.max_column
        lock_column = 0
        for i in range(1, max_column + 1):
            value = table.cell(4, i).value
            if value == 'uiPage':
                lock_column = i

        for i in range(6, max_row + 1):
            value = table.cell(i, lock_column).value
            if value is not None:
                self.mData.append(value)

        # print(self.mData)
        self.write_csharp_code_from_data()
        return True

    def write_csharp_code_from_data(self):
        file_content = "\n"
        file_content += "//-----------------------------------------------\n"
        file_content += "//              生成代码不要修改\n"
        file_content += "//-----------------------------------------------\n"
        file_content += "\n"
        file_content += "using System.Collections.Generic;\n"
        file_content += "using System.IO;\n"
        file_content += "using System.Text;\n"
        file_content += "using UnityEngine;\n"
        file_content += "\n"

        file_content += "public enum UIPage\n"
        file_content += "{\n"
        for index, value in enumerate(self.mData):
            file_content += self.Tab(1) + value + " = " + str(index) + ",\n"
        file_content += "}\n"
        file_content += "\n"

        # save
        path = "UIPage.cs"
        path = UnityCodeDir + path
        path = os.path.splitext(path)[0]
        path = path + ".cs"
        filedir = os.path.dirname(path)
        if not os.path.exists(filedir):
            os.makedirs(filedir)

        file = open(path, "wb")
        file.write(file_content.encode())
        file.close()
